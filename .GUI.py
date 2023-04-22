import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import tkinter as tk
import tkinter.ttk as ttk
import glob
from PIL import Image, ImageTk
import  tkinter.messagebox as msb
from time import strftime
import sqlite3
from openpyxl import load_workbook,Workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from pandas import read_excel, read_csv, concat, DataFrame

#internal modules
from toolbarButtons import ToolbarIcon

def image_configure(widget, path):
    """Accepts widget and path, opens file, creates an instance of image and configures it to the widget"""
    image = ImageTk.PhotoImage(Image.open(path))
    widget.image = image
    widget.configure(image = image)


def pop_up_window(master, text):
    """Accepts parent widget and creates toplevel window with child widgets"""
    window = tk.Toplevel(master, bg = "dark grey")
    window.title(name)
    window.geometry("700x500+270+100")
    window.overrideredirect(1)
    window.grid_rowconfigure(0, weight = 1)
    window.grid_rowconfigure(1, weight = 20)
    window.grid_columnconfigure(0, weight = 1)
    
    def confirm_delete():
        """Prompts and destroys toplevel widget"""
        if msb.askyesno(name, 'Are you sure you want to close this window?'):
            window.destroy()
    window.protocol('WM_DELETE_WINDOW',confirm_delete)
    
    frame0 = tk.Frame(window, bg = "white")
    frame1 = tk.Frame(window, bg = "white")
    frame0.grid(row = 0, column = 0, sticky =tk.N+tk.S+tk.W+tk.E, padx = 2,pady = 1)
    frame1.grid(row = 1, column = 0, sticky =tk.N+tk.S+tk.W+tk.E, padx = 2, pady = 1)
    
    exitButton = tk.Button(frame0, relief=tk.FLAT, command=window.withdraw, bg = "white")
    image_configure(exitButton, 'exit.png')
    exitButton.pack(side = tk.LEFT, padx = 2, pady = 2)
    
    tk.Label(frame0,text = text, relief = tk.FLAT,font = ('Arial',12), anchor = tk.W, bg = "white").pack(side = tk.LEFT,fill = tk.X)
    
    return window, frame1


def color(master):
    """Accepts parent widget and allows configuration of sidebar frame background color"""
    screen, frame = pop_up_window(master, "Set Color")
    frame.grid_columnconfigure(0, weight = 1)
    
    canvas_frame = window_frame(frame, row = 1, column = 0, rowspan = 1, width = 600, height = 443, sticky = tk.N+tk.S+tk.E+tk.W, scroll = 1, anchor = tk.W)
      
    color_var = tk.StringVar()
    color_var.set(open_file("color.txt"))
    
    colors = (("white" , "white"), ("grey" , "grey"), ("orange" , "orange"), ("aliceblue" , "aliceblue"), ("antiquewhite" , "antiquewhite"), ("beige" , "beige"), ("bisque" , "bisque"), ("brown" , "brown"), ("burlywood" , "burlywood"), ("chocolate" , "chocolate"), ("darkgoldenrod" , "darkgoldenrod"), ("darkkhaki" , "darkkhaki"), ("lavender" , "lavender"), ("rosybrown" , "rosybrown"), ("silver" , "silver"))
    
    def _apply_color():
        """Configures background color to the widget at runtime and saves color variable in external file"""
        with open("color.txt","w") as file:
            file.write(color_var.get())
        frame_color_1.configure(bg = open_file("color.txt"))
        last_label.configure(bg = open_file("color.txt"))
    
    [tk.Radiobutton(canvas_frame,text = item[0], fg = item[0], relief = tk.FLAT, value = item[1], font = ("Calibri", 10, "bold"), variable = color_var, command = _apply_color, anchor = tk.W).pack(padx = 30, pady = 10, expand = 1, fill = tk.X) for item in colors]
    
    
def organisation(master):
    """Accepts a toplevel window and saves organisation in external text file"""
    screen, frame = pop_up_window(master, "Organisation Name")
    
    company_name = tk.StringVar()
    company_name.set(name)
    
    def set_organisation_name(variable):
        global name
        name = variable.get()
        tool_bar_label.config(text = name + ' SMS' + strftime(open_file("date.txt")))
        
        with open("organisation_name.txt", "w") as file:
            file.write(name)
    
    tk.Entry(frame, textvariable = company_name, bg = "light grey", relief = tk.FLAT, font = ("Calibri", 13)).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 50)
    ttk.Button(frame, text = "SUBMIT", command= lambda : set_organisation_name(company_name)).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 50)


def graph_style(master):
    """Accepts parent widget, configures theme and saves it in external text file"""
    screen, frame = pop_up_window(master, "Set Graph Style")
    canvas_frame = window_frame(frame, row = 1, column = 0, rowspan = 1, width = 600, height = 443, sticky = tk.N+tk.S+tk.E+tk.W, scroll = 1, anchor = tk.W)
    frame.grid_columnconfigure(0, weight = 1)
    
    style_var = tk.StringVar()
    style_var.set(open_file("graph_style.txt"))
    
    def _apply_style():
        """Configures and saves selected style"""         
        with open("graph_style.txt","w") as file:
            file.write(style_var.get())
                   
    [tk.Radiobutton(canvas_frame,text = item, relief = tk.FLAT, value = item, font = ("Calibri", 10, "bold"), variable = style_var, command = _apply_style, anchor = tk.W).pack(padx = 30, pady = 10, expand = 1, fill = tk.X) for item in plt.style.available]


def theme(master):
    """Accepts parent widget, configures theme and saves it in external text file"""
    screen, frame = pop_up_window(master, "Set Theme")
    
    theme_var = tk.StringVar()
    theme_var.set(style.theme_use())
    
    def _apply_theme():
        """Configures and saves selected theme"""
        style.theme_use(theme_var.get())
        theme_var.set(style.theme_use())      
        with open("theme.txt","w") as file:
            file.write(style.theme_use())
                   
    [tk.Radiobutton(frame,text = item, relief = tk.FLAT, value = item, font = ("Calibri", 10, "bold"), variable = theme_var, command = _apply_theme, anchor = tk.W).pack(padx = 30, pady = 10, expand = 1, fill = tk.X) for item in style.theme_names()]


def Settings():
    screen, frame = toplevel_screen(text = "Settings")
    frame.grid_columnconfigure(0, weight = 1)
    
    options_buttons = {"Select the font color" : lambda : color(screen), "Edit Organisation Name" : lambda : organisation(screen), "Change theme" : lambda : theme(screen),"Change graph style" : lambda : graph_style(screen)}
    
    [tk.Button(frame, text = key, relief = tk.FLAT,bg = 'white', font = ('Arial',10), command = options_buttons[key], anchor = tk.W).pack(expand = 1, fill = tk.BOTH) for key in options_buttons]
    
    show_date = tk.StringVar()
    show_date.set(open_file("date.txt"))
    
    def configure_label():        
        with open("date.txt","w") as file:
            file.write(show_date.get())
            
        tool_bar_label.configure(text = name + ' SMS' + strftime(open_file("date.txt")))
          
    tk.Checkbutton(frame, text = "Show Date", relief = tk.FLAT, bg = "white", onvalue = " ~ %A, %B %d, %G", offvalue = "", height = 1, font = ('Arial', 10), variable = show_date, command = configure_label, anchor = tk.W).pack(expand = 1, fill = tk.BOTH)
    
           
def SignUp():
    """Allows you to save your details in database"""
    screen, frame = toplevel_screen(text = "Sign Up")
    
    font = ("Calibri", 10)
    username_signup = tk.StringVar()
    password_signup = tk.StringVar()
    password_confirm = tk.StringVar()
    variables = [(username_signup,"Enter username"), (password_signup,"Enter password"), (password_confirm, "Re-enter to confirm password")]
    [variable[0].set(variable[1]) for variable in variables]
    
    def Sign_Up_Config():
        """Logic for the sign up function"""
        [variable[0].set(variable[1]) for variable in variables]
    
    label_image = tk.Label(frame, bg = "white", relief = tk.FLAT)
    image_configure(label_image, 'signup_label.png')
    label_image.pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 2)
    
    [tk.Entry(frame, textvariable = variable[0], bg = "light grey",relief = tk.FLAT).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15) for variable in variables]

    ttk.Button(frame, text = "SIGN UP", command= Sign_Up_Config).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15)
    
    label = tk.Label(frame, bg = "white", relief = tk.FLAT, font = ("Calibri", 10), anchor = tk.W)
    label.pack(expand = 1, fill = tk.BOTH, padx = 100, pady = [2, 20])
    
    
def SignIn():
    """Allows you to access your account"""
    screen, frame = toplevel_screen(text = "Sign In")
    
    font = ("Calibri", 10)
    username_signin = tk.StringVar()
    password_signin = tk.StringVar()
    variables = [(username_signin, "Enter username"), (password_signin, "Enter password")]
    [variable[0].set(variable[1]) for variable in variables]
    
    def Sign_In_Config():
        """Logic for the sign in function"""
        [variable[0].set(variable[1]) for variable in variables]
        
    label_image = tk.Label(frame, bg = "white", relief = tk.FLAT)
    image_configure(label_image, 'signin_label.png')
    label_image.pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 2)
    
    [tk.Entry(frame, textvariable = variable[0], bg = "light grey",relief = tk.FLAT).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15) for variable in variables]
    
    tk.Checkbutton(frame, text = "Keep me logged in", relief = tk.FLAT, bg = "white", height = 1, font = ('Arial', 10), variable = None, command = None, anchor = tk.W).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15)
    
    ttk.Button(frame, text = "SIGN IN", command= Sign_In_Config).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15)
    
    label = tk.Label(frame, bg = "white", relief = tk.FLAT, font = ("Calibri", 10), anchor = tk.W)
    label.pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15)
    
    
def GetPremium(master):
    """Accepts parent widget and displays capabilities of various versions of the application"""
    screen, frame = pop_up_window(master, "Get Premium")
    frame.grid_columnconfigure(0, weight = 1)
    
    basic_data = ("Basic Sales Editor", "Basic View of Stock", "Basic Graphical Representation of \nSales Data", "Basic Sales Analysis")
    premium_data = ("Sales Prediction", "Advanced Sales Analysis", "Advanced Graphical Representation \nof Sales Data", "Price Prediction")
    
    tk.Label(frame, text = "Unlock Premium Features", fg = "red", bg = "white", font = ("Calibri", 10), anchor = tk.W).grid(row = 0, column = 0, columnspan = 2, sticky = tk.W+tk.E, pady = 5)
    
    canvas_frame = window_frame(frame, row = 1, column = 0, rowspan = 1, width = 600, height = 360, sticky = tk.N+tk.S+tk.E+tk.W, scroll = 1, anchor = tk.W, bg = "white")
    
    ttk.Button(frame, text = "Purchase Premium", command = None).grid(pady = 4, padx = 30,row = 2, column = 0, columnspan = 2, sticky = tk.E+tk.W)
    
    [tk.Label(canvas_frame, text = basic_data[i], bg = "white", font = ('Calibri', 9), anchor = tk.W, justify = tk.LEFT).grid(row = i + 1, column = 0, sticky = tk.N+tk.S+tk.W+tk.E, pady = 5, padx = 10) for i in range(len(basic_data))]
    
    [tk.Label(canvas_frame, text = premium_data[i], bg = "white", font = ('Calibri', 9), anchor = tk.W, justify = tk.LEFT).grid(row = len(basic_data) + 1 + i, column = 0, sticky = tk.N+tk.S+tk.W+tk.E, pady = 5, padx = 10) for i in range(len(premium_data))]
    
    options = {"sticky" : tk.N+tk.S+tk.W+tk.E, "pady" : 5, "padx" : 10}
    tk.Label(canvas_frame, text = "Premium" ,bg = "white", font = ('Calibri', 9, "bold italic underline")).grid(row = 0, column = 1, **options)
    
    tk.Label(canvas_frame, text = "Basic", bg = "white", font = ('Calibri', 9, "bold italic underline")).grid(row = 0, column = 2, **options)
    
    [tk.Label(canvas_frame, text = "√" ,bg = "white", font = ('Calibri', 9, "bold"), fg = "green").grid(row = i + 1, column = 1, **options) for i in range((len(premium_data)+len(basic_data)))]
    
    [tk.Label(canvas_frame, text = "√" ,bg = "white", font = ('Calibri', 9, "bold"), fg = "green").grid(row = i + 1, column = 2, **options) for i in range(len(basic_data))]
    
    [tk.Label(canvas_frame, text = "x" ,bg = "white", font = ('Calibri', 9, "bold"), fg = "red").grid(row = i + 1 + len(basic_data), column = 2, sticky = tk.N+tk.S+tk.W+tk.E, pady = 5, padx = 10) for i in range(len(premium_data))]
   
   
def PrivacyPolicy(master):
    screen, frame = pop_up_window(master, "Privacy Policy")
    frame.grid_columnconfigure(0, weight = 1)
    
    canvas_frame = window_frame(frame, row = 1, column = 0, rowspan = 1, width = 600, height = 443, sticky = tk.N+tk.S+tk.E+tk.W, scroll = 1, anchor = tk.W, bg = "whitesmoke")
    
    data = ("This is a product of Our Digital Times and should not be used whatsover for any other purpose other than the intended one. \nAs a user of this product you are urged to share  substantial feedback to our developers so as to improve our services. The feedback will be used in analysis projects and privacy is upheld in the process. \nWe urge you not to disclose any private or sensitive information while sending comments, complaints or feedback , even when prompted by our contacts, so as to uphold your constitutional right to privacy.\nIn case of piracy of our products and services, kindly inform us through our media handles and contacts - this helps us provide you with better services.")
    
    tk.Label(canvas_frame, text = "Terms and Conditions", bg = "whitesmoke", fg = "red", font = ('Calibri', 9, "bold")).grid(row = 0, column = 0, sticky = tk.N+tk.S+tk.W+tk.E, pady = 7, padx = 2)
    
    text_widget = tk.Text(canvas_frame, bg = "whitesmoke", relief = tk.FLAT, font = ('Calibri', 9))
    text_widget.insert(tk.END, data)
    text_widget.config(width = 50, height = 10, wrap = tk.WORD, spacing1 = 1, spacing2 = 2, spacing3 = 20, state = tk.DISABLED)
    text_widget.grid(row = 1, column = 0, sticky = tk.N+tk.S+tk.W+tk.E, pady = 7, padx = 2)
    
    ttk.Button(canvas_frame, text = "ACCEPT", command = None).grid(row = 2, column = 0, sticky = tk.E+tk.W, pady = 4, padx = 30)

                
def AboutUs(master):
    screen, frame = pop_up_window(master, "About Us")
    frame.grid_columnconfigure(0, weight = 1)
    
    canvas_frame = window_frame(frame, row = 1, column = 0, rowspan = 1, width = 600, height = 443, sticky = tk.N+tk.S+tk.E+tk.W, scroll = 1, anchor = tk.W, bg = "black")
    
    data = ("Developer Organisation: Our Digital Times Organisation", "Names : Barasa Michael Murunga", "Year of Production : 2020", "Releases : 2018, 2020", "Mission : To enhance economic productivity through \ntechnology.", "Location : Mariakani, Kenya","Email Address : ourdigitaltimes@gmail.com", "Website : www.ourdigitaltimes.co.ke", "Facebook : ourdigitaltimeske", "Instagram : ourdigitaltimeske", "Mobile No. : 0793770236")
    
    images = glob.glob('*.gif')
    def slideshow(index = 0):
        image_configure(label_image, images[index])
        index += 1
        if index == len(images) : index = 0
        label_image.after(3000, lambda : slideshow(index))
        
    label_image = tk.Label(canvas_frame, bg = "black")
    label_image.grid(row = 0, column = 0, pady = 2)
    
    [tk.Label(canvas_frame, text = data[i], bg = "black", fg = "green", font = ('Calibri', 9), anchor = tk.W).grid(row = i + 1, column = 0, sticky = tk.N+tk.S+tk.W+tk.E, pady = 7, padx = 2) for i in range(len(data))]
    
    slideshow()
    
    
def help_email(email_address, msg, status_label):
    try:
        mail_content = msg.get()    
        
        sender_address = 'ourdigitaltimes@gmail.com'
        sender_password = 'PCRdh@12795'
        receiver_address = 'mikebarasa03@gmail.com'
             
        message = MIMEMultipart()
        message['From'] = sender_address
        message['To'] = receiver_address
        message['Subject'] = 'Help Request for %s (personal email - %s):'% (open_file("organisation_name.txt"), email_address)
        
        message.attach(MIMEText(mail_content, 'plain'))
        
        session = smtplib.SMTP('smtp.gmail.com', 587)
        session.starttls()
        session.login(sender_address, sender_password)
        
        text = message.as_string()
        session.sendmail(sender_address, receiver_address, text)
        session.quit()
        
        status_label.configure(text = 'Message sent successfully', fg = 'green')
        email_address.set("Enter email address")
        msg.set("Type Message")
    except:
        status_label.configure(text = 'Attempt to send message was unsuccessful. Check your internet connection and try again.', fg = 'red')


def Help():
    screen, frame= toplevel_screen(text = "Help")
    
    font = ("Calibri", 10)
    message = tk.StringVar()
    message.set("Type Message")
    email_address = tk.StringVar()
    email_address.set("Enter email address")
     
    [tk.Label(frame,text = item, bg = "white", relief = tk.FLAT, font = font, anchor = tk.W).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 2) for item in ["Version : 1.0.0", "Email us for assistance "]]
    
    status_label = tk.Label(frame, bg = "white", relief = tk.FLAT, font = font, anchor = tk.W)
    status_label.pack(side = tk.BOTTOM, expand = 1, fill = tk.BOTH, padx = 100, pady = 2)
    
    [tk.Entry(frame, textvariable = item, bg = "light grey",relief = tk.FLAT, font = font).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 10) for item in [message, email_address]]
    
    ttk.Button(frame, text = "Send email", command= lambda : help_email(email_address, message, status_label)).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 15)
    tk.Button(frame, text = "About Us", bg = "white",relief = tk.FLAT, font = font, command = lambda : AboutUs(screen), anchor = tk.W).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 2)
    tk.Button(frame, text = "Privacy Policy", bg = "white", relief = tk.FLAT, font = font, command = lambda : PrivacyPolicy(screen), anchor = tk.W).pack(expand = 1, fill = tk.BOTH, padx = 100, pady = 2)
    
        
def toplevel_screen(text =None):
    """Accepts label text and creates child widgets"""
    screen = tk.Toplevel()
    screen.attributes("-fullscreen",True)    
    screen.grid_rowconfigure(0, weight = 1)
    screen.grid_rowconfigure(1, weight = 20)
    screen.grid_columnconfigure(0, weight = 1)
    
    def confirm_delete():
        """Prompts and destroys toplevel window"""
        if msb.askyesno(name, 'Are you sure you want to close this window?'):
            screen.destroy()
    screen.protocol('WM_DELETE_WINDOW',confirm_delete)
    
    frame0 = tk.Frame(screen)
    frame1 = tk.Frame(screen, bg = "white")
    frame0.grid(row = 0, column = 0, sticky =tk.N+tk.S+tk.W+tk.E)
    frame1.grid(row = 1, column = 0, sticky =tk.N+tk.S+tk.W+tk.E)
    
    exitButton = tk.Button(frame0, relief=tk.FLAT, command=screen.withdraw)
    image_configure(exitButton, 'exit.png')
    exitButton.pack(side = tk.LEFT, padx = 2, pady = 2)
    
    tk.Label(frame0,text = text, relief = tk.FLAT,font = ('Arial',12), anchor = tk.W).pack(side = tk.LEFT,fill = tk.X)
    
    return screen, frame1


def month_totals(filename):
    """Accepts path to excel file and calculates the monthly totals"""
    data = [read_excel(filename, sheet_name = sheet.title, usecols = [5]) for sheet in load_workbook(filename)]
    summed = concat(data, axis = 1, keys = [sheet.title for sheet in load_workbook(filename)]).sum().reset_index(drop = True, level = 1)
    dataframe = summed.reset_index()
    dataframe.columns = ['Month','Total']
    
    return dataframe
    
def attained_profit():
    filename = excel()
    sheetnames = [sheet.title for sheet in load_workbook(filename)]
        
    dict = {}
    for i in range(len(sheetnames)):
        dataframe = read_excel(filename, sheet_name = sheetnames[i], usecols = 'D,F,G')
        dataframe['Profit'] = dataframe.Total - (dataframe.Quantity * dataframe['Cost Price'])
        dict[sheetnames[i]] = int(dataframe['Profit'].sum())
        
    frame = DataFrame({'Month' : list(dict.keys()), 'Profit' : list(dict.values())})
        
    return frame


def summary(master):
    """Accepts parent widget, analyses data from all persistent storage mediums and displays a summary of it"""
    ttk.Label(master, text = "Analysed Sales  and Stock Records", font = ("Calibri",12,"bold underline italic")).grid(row = 0, column = 0, padx = 5)
    
    canvas_frame = window_frame(master, row = 1, column = 0, rowspan = 1, width = 800, height = 450, sticky = tk.N+tk.S+tk.E+tk.W, pady = 20, scroll = 1, anchor = tk.W)
      
    tk.Label(canvas_frame, text = "Sold Stock Net Value",fg = "blue", font = ("Calibri",10,"bold underline")).grid(row = 0, column = 0, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W, columnspan = 3)
    tk.Label(canvas_frame, text = "Monthly Records : ", font = ("Calibri",10,"bold"),anchor = tk.W).grid(row = 1, column = 1, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)
    
    #display headers
    data = month_totals(excel())
    col_num = 1
    for col in data.columns:
        tk.Label(canvas_frame, text = col, fg = "red", font = ("Arial",9), anchor = tk.W).grid(padx = 7, pady = 3, row=2, column = col_num)
        col_num += 1    
    #display data values
    r = 3
    for row in data.values:
        c = 1
        for cell in row:
            tk.Label(canvas_frame,text=cell, anchor = tk.W).grid(padx = 7, pady = 3, row=r,column = c)
            c+=1
        r+=1
    
    tk.Label(canvas_frame, text = "Total Sales :", font = ("Arial",10,"bold"), anchor = tk.W).grid(row = r + 1, column = 1, padx = 5)
    
    tk.Label(canvas_frame, text = data['Total'].sum(), font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 1, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)    
    #Retrieving remaining stock from database
    cursor,conn = Database()   
    try:
        cursor.execute("SELECT SUM(Quantity*Selling_Price) FROM `stock " + year + "`")
        fetch = cursor.fetchall()
        gross_stock = fetch[0]
        conn.commit()
    except:
        msb.showerror(name,"An error occured while retrieving requested data!!!")
    cursor.close()
    conn.close()    
    #display widgets
    tk.Label(canvas_frame, text = "Remaining Stock Net Value",fg = "blue", font = ("Arial",10,"bold underline")).grid(row = r + 2, column = 1, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W, columnspan = 3)
    tk.Label(canvas_frame, text = "Remaining Stock : ", font = ("Arial",10,"bold"), anchor = tk.W).grid(row = r + 3, column = 1, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)    
    tk.Label(canvas_frame, text = gross_stock[0], font = ("Arial",10,"bold"), anchor = tk.W).grid(row = r + 3, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)    
    #Gross Profit    
    cursor,conn = Database()
    
    try:
        #selects all names of tables in master_table
        cursor.execute("SELECT name FROM sqlite_master WHERE type ='table' AND name LIKE ?", ('%stock-original%',))
        fetch = cursor.fetchall()
        table_name_tuples = [fetch[i] for i in range(len(fetch))]
        net_stock = 0
        net_stock1 = 0
        for i in range(len(table_name_tuples)):
            table = table_name_tuples[i]
            cursor.execute("SELECT SUM(Selling_Price*Quantity) FROM `"+table[0]+"`")
            fetch = cursor.fetchall()
            gross_total1 = fetch[0]
            net_stock1 += gross_total1[0]
            cursor.execute("SELECT SUM(Buying_Price*Quantity) FROM `"+table[0]+"`")
            fetch = cursor.fetchall()
            gross_total = fetch[0]
            net_stock += gross_total[0]
            
        conn.commit()
    except:
        msb.showerror(name,"An error occured while retrieving data from stock records.\nNet stock and gross total may not be displayed in the sales summary section!!!")
    cursor.close()
    conn.close()
    
    tk.Label(canvas_frame, text = "Net Stock :", font = ("Arial",10,"bold"), anchor = tk.W).grid(row = r + 4, column = 1, padx = 5)
    
    tk.Label(canvas_frame, text = "Not Implemented", fg = "red", font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 4, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)
    
    tk.Label(canvas_frame, text = "Proposed Profit :", font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 5, column = 1, padx = 5)
    
    tk.Label(canvas_frame, text = int(net_stock1 - net_stock), font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 5, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)    
    #petty cash and expenses
    data_csv = read_csv("Petty Cash "+year + ".txt", names = ['Date','Time','Description','Amount'])  
    tk.Label(canvas_frame, text = "Petty Cash & Expenses :", font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 6, column = 1, padx = 5)  
    tk.Label(canvas_frame, text = data_csv['Amount'].sum(), font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 6, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)  
    
    try:
        frame = attained_profit()
        #display headers
        col_num = 1
        for col in frame.columns:
            tk.Label(canvas_frame, text = col, fg = "red", font = ("Arial",9), anchor = tk.W).grid(padx = 7, pady = 3, row=r + 7, column = col_num)
            col_num += 1
        
        #display data values
        r = r + 8
        for row in frame.values:
            c = 1
            for cell in row:
                ttk.Label(canvas_frame,text=cell, anchor = tk.W).grid(padx = 7, pady = 3, row=r,column = c)
                c+=1
            r += 1
     
        tk.Label(canvas_frame, text = "Attained Net Profit :", font = ("Arial",10,"bold"),anchor = tk.W).grid(row = r + 1, column = 1, padx = 5)
        
        tk.Label(canvas_frame, text = frame['Profit'].sum(), font = ("Arial",10,"bold")).grid(row = r + 1, column = 2, padx = 5, sticky = tk.N+tk.S+tk.E+tk.W)
    except AttributeError:
        msb.showerror(name, "An error occured while retrieving required data! Profit details are going to be momentarily omitted.")

#graphical representations of data
def plots(master):
    filename = excel()
    sheetnames = [sheet.title for sheet in load_workbook(filename)] 
   
    ttk.Label(master, text = "Graphical Representation of Sales", font = ("Calibri",12,"bold underline italic")).grid(row = 0, column = 0, padx = 5)

    def on_configure(event):
        canvas.configure(scrollregion = canvas.bbox('all'))
    
    canvas = tk.Canvas(master,width = 800,height =470, bg = "white")
    canvas.grid(row = 1, column = 0,sticky = tk.N+tk.S+tk.E+tk.W,pady = 20)
    master.grid_columnconfigure(0, weight = 1)
    
    scrollbar = ttk.Scrollbar(master,command = canvas.yview)
    scrollbar.grid(row = 1, column = 1,sticky = tk.N+tk.S,pady = 20, rowspan = 1)
    canvas.configure(yscrollcommand = scrollbar.set)
    
    canvas.bind('<Configure>',on_configure)
    
    canvas_frame = tk.Frame(canvas)
    canvas.create_window((0,0),window = canvas_frame, width = 812)
    try:
        for sheetname in sheetnames:
            data = read_excel(filename,sheet_name = sheetname)
            x = data['Quantity'].groupby(data['Item']).sum().sort_values(ascending = False).head(10)
            plt.rcParams.update({'figure.autolayout' : True})
            fig = plt.figure(figsize = (8,6))
            plt.style.use(open_file("graph_style.txt"))            
            ax = plt.gca()
            try:
                x.plot.bar()
            except:
                ax.bar([], [])
            plt.title('Top 10 Sold Products ' + sheetname,fontsize = 16,color = 'red',fontweight = 'bold')
            plt.xlabel('Products',fontsize = 14)
            plt.ylabel('Quantity sold',fontsize = 14)         
            ax.spines['top'].set_color('none')
            ax.spines['right'].set_color('none')          
            
            canvas1 = FigureCanvasTkAgg(fig,canvas_frame)
            canvas1.draw()
            canvas1.get_tk_widget().pack(fill = tk.BOTH, expand = True)
            canvas1._tkcanvas.pack(fill = tk.BOTH, expand = True)
    except SyntaxError:
        msb.showerror(name, "An error occured while retrieving data!")
    
    #pie chart
    try:
        fig1 = plt.figure()
        pie_data_frame = month_totals(excel())
        pie_columns_x = pie_data_frame['Month']
        pie_rows_y = pie_data_frame['Total']
        plt.title('Monthly Sales - Pie Chart',fontsize = 16,color = 'red',fontweight = 'bold')
        plt.pie(pie_rows_y, labels = pie_columns_x, autopct = '%1.1f%%')
        canvas1 = FigureCanvasTkAgg(fig1,canvas_frame)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill = tk.BOTH, expand = True)
        canvas1._tkcanvas.pack(fill = tk.BOTH, expand = True)
        
        fig2 = plt.figure()        
        plt.title('Monthly Sales - Line Graph',fontsize = 16,color = 'red',fontweight = 'bold')
        plt.plot(pie_columns_x, pie_rows_y)
        plt.xlabel('Months',fontsize = 14)
        plt.ylabel('Total in Kshs.',fontsize = 14)
        ax = plt.gca()
        ax.spines['top'].set_color('none')
        ax.spines['right'].set_color('none')
        canvas2 = FigureCanvasTkAgg(fig2,canvas_frame)
        canvas2.draw()
        canvas2.get_tk_widget().pack(fill = tk.BOTH, expand = True)
        canvas2._tkcanvas.pack(fill = tk.BOTH, expand = True)
        
        pie_data_frame = attained_profit()
        pie_columns_x = pie_data_frame['Month']
        pie_rows_y = pie_data_frame['Profit']
        fig3 = plt.figure()        
        plt.title('Monthly Profit - Line Graph',fontsize = 16,color = 'red',fontweight = 'bold')
        plt.plot(pie_columns_x, pie_rows_y)
        plt.xlabel('Months',fontsize = 14)
        plt.ylabel('Profit in Kshs.',fontsize = 14)
        ax = plt.gca()
        ax.spines['top'].set_color('none')
        ax.spines['right'].set_color('none')
        canvas2 = FigureCanvasTkAgg(fig3,canvas_frame)
        canvas2.draw()
        canvas2.get_tk_widget().pack(fill = tk.BOTH, expand = True)
        canvas2._tkcanvas.pack(fill = tk.BOTH, expand = True)      
    except SyntaxError:
        msb.showerror(name, "An error occured while retrieving required data!")


def showSales(master):
    """Accepts parent widget and displays excel file data in tabular form"""
    file = excel()
    wb1 = load_workbook(file, data_only=True)
    
    def display(sheet):
        r = 0
        try:
            for row in ws1:
                c = 0
                for cell in row:
                    ttk.Label(canvas_frame,text=cell.value).grid(padx = 4, pady = 3, row=r,column=c)
                    c+=1
                r+=1
        except TypeError:
            msb.showerror(name, "The requested worksheet is empty. No data will be displayed.")
    
    def refresh(*args):
        """Refreshes displayed records based on chosen sheetname"""
        ws1 = wb1[month_var.get()]
        label_.config(text = "Sales Records ~ "+month_var.get())
        for child in canvas_frame.winfo_children():
            child.destroy()
        
        display(ws1)
        
    month_var = tk.StringVar()
    month_var.trace("w", refresh)
    
    ws1 = wb1[month]
    values = [sheet.title for sheet in wb1]
    
    label_ = tk.Label(master, text = "Sales Records ", font = ("Calibri",12,"bold underline italic"))
    label_.grid(row = 0, column = 0, padx = 5)
    combo_box = ttk.Combobox(master, values = values, textvariable = month_var)
    combo_box.grid(row = 1, column = 0, sticky = tk.W )
    
    canvas_frame = window_frame(master, row = 2, column = 0, rowspan = 1, width = 800, height = 420, sticky = tk.N+tk.S+tk.E+tk.W, pady = 20, padx = 0, scroll = 1, anchor = tk.W)
    
    display(ws1)


def read_txt(master, text, filename = None, names = None):
    """Accepts parent widget, filename and columns then displays the data in tabular form"""
    ttk.Label(master, text = text, font = ("Calibri",12,"bold underline italic")).grid(row = 0, column = 0, padx = 5)
    
    canvas_frame = window_frame(master, row = 1, column = 0, rowspan = 1, width = 800, height = 470, sticky = tk.N+tk.S+tk.E+tk.W, pady = 20, padx = 0, scroll = 1, anchor = tk.W)
    try:
        data = read_csv(filename, names = names)
        
        col_num = 0
        for col in data.columns:
            tk.Label(canvas_frame, text = col, fg = "red", font = ("Arial",9)).grid(padx = 10, pady = 3, row=0, column = col_num)
            col_num += 1
        r = 1
        for row in data.values:
            c = 0
            for cell in row:
                ttk.Label(canvas_frame,text=cell).grid(padx = 10, pady = 3, row=r,column = c)
                c+=1
            r+=1
    except:
        msb.showerror(name, "Couldn't retrieve data from {} file!\nFile is probably empty.".format(text))


def update_stock(change):
    """Updates the stock database"""    
    cursor,conn = Database()
    
    try:
        cursor.execute("UPDATE `stock " + year +"` SET Quantity = ? WHERE Product_ID = ?", (change,prod_id))
        conn.commit()
    except:
        msb.showerror(name,"An error occured while updating the stock!!!")
    
    try:
        cursor.execute("SELECT Product, Buying_Price FROM `stock " + year +"` WHERE `Quantity` <= 0")
        fetch = cursor.fetchall()
        list1 = [i for i in fetch[0]]
        with open(month + ".txt",'a') as file:
            file.write(date + ',' + list1[0] + ',' + list1[1] + '\n')
    except IndexError:
        pass
    
    cursor.execute("DELETE FROM `stock " + year +"` WHERE `Quantity` <= 0")
    conn.commit()
    cursor.close()
    conn.close()


def savePettyCash():
    """Saves petty cash data on csv format file"""
    if petty_amount.get() > 0 and petty_description.get() != "":
        with open("Petty Cash "+year + ".txt",'a') as file:
                file.write(strftime("%x")+','+strftime('%H:%M:%S %p')+','+petty_description.get()+','+str(petty_amount.get())+'\n')
        
        msb.showinfo(name,"Petty Cash Record Saved!")
    
    else:
        msb.showerror(name,"Invalid data!\nTo troubleshoot, ensure:\n\t1. Expenditure Description is not empty\n\t2. Amount is greater than 0")
    
    petty_description.set('')
    petty_amount.set(0.0)


def saveSalesData():
        """Saves sales data in excel file and clears entry widgets"""
        filename = excel()
        wb = load_workbook(filename)
        sheet = wb[month]
        
        first_column = sheet['A']
        second_column = sheet['B']
        third_column = sheet['C']
        fourth_column = sheet['D']
        fifth_column = sheet['E']
        sixth_column = sheet['F']
        seventh_column = sheet['G']
        
        col_len1 = str(len(first_column)+1)
        col_len2 = str(len(second_column)+1)
        col_len3 = str(len(third_column)+1)
        col_len4 = str(len(fourth_column)+1)
        col_len5 = str(len(fifth_column)+1)
        col_len6 = str(len(sixth_column)+1)
        col_len7 = str(len(seventh_column)+1)
              
        if salesquantity.get() > 0 and _Cost.get() > 0:    
            product = salesproduct.get()
            price = salesSP.get()
            quantity = salesquantity.get()
            if quantity <= int(amount) :
                cost = _Cost.get()
        
                total = 0
                for i in range(quantity):
                    total += float(price)
                
                sheet['A' + col_len1] = date
                sheet['B' + col_len2] = strftime("%H:%M:%S %p")
                sheet['C'+ col_len3] =  product
                sheet['D' + col_len4] = quantity
                sheet['E' + col_len5] =  price
                sheet['F' + col_len6] =  total
                sheet['G' + col_len7] =  cost
                wb.save(filename)
                
                change = int(amount) - quantity
                update_stock(change)
            
            else:
                msb.showerror(name, "There is lesser stock of the product in the database\n than the quantity requested")       
         
        else:
            msb.showerror(name,'Invalid data submitted!\nTo troubleshoot :\n1. Enter valid value in quantity field\n2. Select item from list rather than typing it.')
            
        Reset()
        salesproduct.set("")
        salesSP.set(0.0)
        salesquantity.set(0)
        _Cost.set(0)


def select():
    """Inserts selected data from tree_widget into their respective widgets for persistent entry into excel file"""
    global amount, prod_id
    
    if not salestree.selection():
       msb.showinfo(name,'Kindly make a selection !')
    else:
        curItem = salestree.focus()
        contents =(salestree.item(curItem))
        selecteditem = contents['values']
        cursor, conn = Database()
        
        try:
            cursor.execute("SELECT Product, Product_ID, Quantity, Selling_Price, Buying_Price FROM `Stock " + year +"` WHERE `Product_ID` = %d" % selecteditem[0])
            fetch = cursor.fetchall()
            iproduct, prod_id, amount, price, cost = (value for value in fetch[0])
            salesproduct.set(iproduct)
            salesSP.set(price)
            _Cost.set(cost)
            
        except SyntaxError:
            msb.showerror(name,'An error occured!!!')
        
        conn.commit()
        cursor.close()
        conn.close()
 
               
def excel():
    """Creates excel file and associated sheets if don't exist and returns filename"""
    filename = year + ".xlsx"
    
    def create_column_headers(wb):
        sheet = wb[month]
        if (sheet['A1'].value == 'Date') and (sheet['B1'].value == 'Time') and (sheet['C1'].value == 'Item') and (sheet['D1'].value == 'Quantity') and (sheet['E1'].value == 'Selling Price') and (sheet['F1'].value == 'Total') and (sheet['G1'].value == 'Cost Price'):
               pass 
        else:
               sheet['A1'] = 'Date'
               sheet['B1'] = 'Time'
               sheet['C1'] = 'Item'
               sheet['D1'] = 'Quantity'
               sheet['E1'] = 'Selling Price'
               sheet['F1'] = 'Total'
               sheet['G1'] = 'Cost Price'
            
    if os.path.isfile(filename):       
       wb = load_workbook(filename)    
       if month in wb.sheetnames:
           pass
       else:         
           wb.create_sheet(title = month)
           create_column_headers(wb)
           wb.save(filename)
    
    else:
       wb = Workbook()
       for sheet in wb.sheetnames:
           sheet = wb.get_sheet_by_name(sheet)
           wb.remove_sheet(sheet)
       wb.create_sheet(title=month)
       create_column_headers(wb)
       wb.save(filename) 
    
    return filename


def tree_widget(master, height = 0):
    """Creates tree widget and returns it to invoking function"""
    scrollbary = ttk.Scrollbar(master, orient=tk.VERTICAL)
    tree = ttk.Treeview(master, columns=("ProductID", "Product Name", "Product Qty", "Product Price"), selectmode="extended", height=height,yscrollcommand=scrollbary.set)
    scrollbary.config(command=tree.yview)
    scrollbary.pack(side=tk.RIGHT, fill=tk.Y)
    tree.heading('ProductID', text="ID",anchor=tk.W)
    tree.heading('Product Name', text="Product Name",anchor=tk.W)
    tree.heading('Product Qty', text="Qty",anchor=tk.W)
    tree.heading('Product Price', text="Price",anchor=tk.W)
    tree.column('#0', stretch=tk.NO, minwidth=0, width=0)
    tree.column('#1', stretch=tk.NO, minwidth=0, width=30)
    tree.column('#2', stretch=tk.NO, minwidth=0, width=200)
    tree.column('#3', stretch=tk.NO, minwidth=0, width=80)
    tree.column('#4', stretch=tk.NO, minwidth=0, width=80)
    tree.pack()
    
    return tree


def SalesExcel(master):
    """Creates relevant widgets for input in the parent parameter accepted"""
    global salestree, salesproduct, salesquantity, salesSP, petty_description, petty_amount, _Cost
    
    salesproduct = tk.StringVar()
    salesquantity = tk.IntVar()
    salesSP = tk.DoubleVar()
    _Cost = tk.DoubleVar()
    petty_description = tk.StringVar()
    petty_amount = tk.DoubleVar()
    
    labels = ["Sales...","Product","Quantity","Selling Price"]
    [ttk.Label(master,text = labels[i]).grid(row = i, column = 0,sticky = tk.N+tk.S+tk.E+tk.W,pady = 5, padx = 10) for i in range(len(labels))]
    
    entry_vars = [salesproduct,salesquantity,salesSP]
    [ttk.Entry(master,textvariable = entry_vars[i]).grid(row = i + 1, column = 1,sticky = tk.N+tk.S+tk.E+tk.W,pady = 5, padx = 10) for i in range(len(entry_vars))]
    
    ttk.Button(master,text = 'SUBMIT',command = saveSalesData).grid(row = 4,column = 1,pady = 5, padx = 10, sticky = tk.E )
    ttk.Separator(master).grid(row=5,column=0,columnspan=2,ipadx = 100,sticky = tk.E+tk.W,pady = 5)
    
    ttk.Label(master,text = 'Petty Cash...').grid(row = 6,column = 0,pady = 2, padx = 10 , sticky = tk.W+tk.E,columnspan = 2)
    ttk.Label(master,text = 'Description :').grid(row = 7,column = 0,pady = 5, padx = 10, sticky = tk.W+tk.E)
    ttk.Label(master,text = 'Amount :').grid(row = 8,column = 0,pady = 5, padx = 10, sticky = tk.W+tk.E)
    ttk.Entry(master,textvariable = petty_description).grid(row = 7, column = 1,pady = 5, padx = 10, sticky = tk.W+tk.E)
    ttk.Entry(master,textvariable = petty_amount).grid(row = 8, column = 1,pady = 5, padx = 10 , sticky = tk.W+tk.E)
    ttk.Button(master,text = 'SUBMIT',command = savePettyCash).grid(row = 9, column = 1,pady = 5, padx = 10, sticky = tk.E)
    ttk.Separator(master).grid(row=10,column=0,columnspan=2,ipadx = 100,sticky = tk.E+tk.W,pady = 5)
    
    ttk.Entry(master,textvariable = SEARCH,width = 32).grid(row = 11,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    ttk.Button(master,text = 'Search',width = 32,command = Search).grid(row = 12,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    ttk.Button(master,text = 'Reset',width = 32,command = Reset).grid(row = 13,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    
    canvas_frame = window_frame(master, row = 0, column = 3, width = 420, height = 100, anchor = tk.W, rowspan = 14)
    salestree = tree_widget(canvas_frame, height = 16)
    DisplayData(salestree)
    
    ttk.Button(master,text = 'SELECT',command = select).grid(row = 4,column = 0,pady = 5, padx = 10 )

         
def update(master):
    """Invokes tree_widget and search, delete and reset buttons """
    global stocktree,SEARCH
    SEARCH = tk.StringVar()
    
    ttk.Label(master,text = "Search...").grid(row = 7,column = 0,columnspan = 2,padx = 10,sticky = tk.W)
    ttk.Entry(master,textvariable = SEARCH,width = 30).grid(row = 8,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    ttk.Button(master,text = 'Search',width = 30,command = Search).grid(row = 9,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    ttk.Button(master,text = 'Reset',width = 30,command = Reset).grid(row = 10,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    ttk.Button(master,text = 'Delete',width = 30,command = Delete).grid(row = 11,column = 0,columnspan = 2,pady = 5, padx = 10,sticky = tk.W)
    
    ttk.Separator(master).grid(row=12,column=0,columnspan=5,ipadx = 100,sticky = tk.E+tk.W,pady = 20)
    
    canvas_frame = window_frame(master, row = 0, column = 3, width = 420, height = 100, anchor = tk.W, rowspan = 12)
    
    stocktree = tree_widget(canvas_frame, height = 15)
    DisplayData(stocktree)


def add_stock(product,quantity,price,selling_price):
        """Accepts data variables and saves them in a database"""    
        cursor,conn = Database()
        
        try:
            cursor.execute("INSERT INTO `stock " + year +"` (Time, Date,Product,Buying_Price,Quantity,Selling_Price) VALUES(?, ?, ?, ?, ?, ?)", (strftime("%H:%M:%S %p"),date,product.title(),price,quantity,selling_price))
            cursor.execute("INSERT INTO `stock-original " + month +"` (Time, Date,Product,Buying_Price,Quantity,Selling_Price) VALUES(?, ?, ?, ?, ?, ?)", (strftime("%H:%M:%S %p"),date,product.title(),price,quantity,selling_price))
            
            conn.commit()
            logfile(3, ' stock database')
        except SyntaxError:
            msb.showwarning(name,"An error occurred while saving the data in the database.")
        
        cursor.close()
        conn.close()


def Database():
    """Creates a database if it doesn't exist alongside its tables and returns the cursor and connection"""
    conn = sqlite3.connect(year + ".db")
    cursor = conn.cursor()
    
    cursor.execute("CREATE TABLE IF NOT EXISTS `stock " + year + "` (Time TEXT, Date TEXT ,Product_ID INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, Product TEXT, Buying_Price TEXT, Selling_Price TEXT, Quantity TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS `stock-original " + month + "` (Time TEXT, Date TEXT ,Product_ID INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, Product TEXT, Buying_Price TEXT, Selling_Price TEXT, Quantity TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS `administrators " + year + "` (ID_Number INTEGER PRIMARY KEY, Names TEXT, Nationality TEXT, Gender TEXT, DOB TEXT, Email_address TEXT, Telephone TEXT, Residence TEXT, Rank TEXT, Password TEXT, Username TEXT)")
    
    return cursor,conn


def addStock(master):
    """Accepts a parent widget and manages widgets on it"""
    stockItem = tk.StringVar()
    stockQuantity = tk.IntVar()
    stockBP = tk.DoubleVar()
    stockSP = tk.DoubleVar()
    
    def submit():
        """Validates data submited"""
        try:
            item = stockItem.get()
            quantity = stockQuantity.get()
            BP = stockBP.get()
            SP = stockSP.get()
        
            if SP > 0 and BP > 0 and quantity > 0 and item != "" and SP > BP:
                add_stock(item,quantity,BP,SP)  
        except:
            msb.showerror(name,"Invalid data!\nTo troubleshoot, ensure:\n\t1. Product Name is not empty\n\t2. Quantity, Selling price and Buying price are greater than 0\n\t3. Selling price is greater than Buying price")
                  
        Reset()
        stockItem.set("")
        stockQuantity.set(0)
        stockBP.set(0.0)
        stockSP.set(0.0)
    
    label_texts = ["ADD STOCK :","Product","Quantity","Buying Price","Selling Price"]
    [ttk.Label(master,text = label_texts[i]).grid(row = i, column = 0,sticky = tk.N+tk.S+tk.E+tk.W,pady = 5, padx = 10) for i in range(len(label_texts))]
    
    text_variables = [stockItem,stockQuantity,stockBP,stockSP]
    [tk.Entry(master,textvariable = text_variables[i]).grid(row = i + 1, column = 1,sticky = tk.N+tk.S+tk.E+tk.W,pady = 5, padx = 10) for i in range(len(text_variables))]
    
    ttk.Button(master,text = 'SUBMIT',command = submit).grid(row = 5,column = 1,pady = 5, padx = 10 )
    ttk.Separator(master).grid(row=6,column=0,columnspan=2,ipadx = 100,sticky = tk.E+tk.W,pady = 20)


def exit_button():
    """Prompts for and destroys root window appropriately"""
    if msb.askyesno(name, "Are you sure you want to exit? "):
        logfile(5)
        exit()


def Reset():
        """Resets tree widgets"""
        stocktree.delete(*stocktree.get_children())
        DisplayData(stocktree)
        salestree.delete(*salestree.get_children())
        DisplayData(salestree)
        SEARCH.set("")


def Search():
    """Search data in database and display it based on tree widget."""
    if SEARCH.get() != "":
        stocktree.delete(*stocktree.get_children())
        salestree.delete(*salestree.get_children())
        
        cursor, conn = Database()
        cursor.execute("SELECT Product_ID, Product,Quantity,Selling_Price FROM `stock " + year + "` WHERE `Product` LIKE ?", ('%'+str(SEARCH.get())+'%',))
        
        fetch = cursor.fetchall()
        for data in fetch:
            stocktree.insert('', 'end', values=(data))
            salestree.insert('', 'end', values=(data))
        
        cursor.close()
        conn.close()


def Delete():
    """Deletes selected item from the stock database"""
    
    if stocktree.selection():
        if msb.askquestion(name, 'Are you sure you want to delete this record?', icon="warning") == 'yes':
            curItem = stocktree.focus()
            contents =(stocktree.item(curItem))
            selecteditem = contents['values']
            stocktree.delete(curItem)
            salestree.delete(curItem)
            cursor, conn = Database()
            cursor.execute("DELETE FROM `stock " + year + "` WHERE `Product_ID` = %d" % selecteditem[0])
            conn.commit()
            cursor.close()
            conn.close()
   
    else:
        msb.showinfo(name,'Kindly make a selection !')
        

def logfile(t,x = 0):
    """Updates the log file for reference purposes"""
    with open('MOSMS.log','a') as file:
        if t == 1:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : [Program initiated]'+ ';\n')
        if t == 2:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Access level '+ x+ ';\n')
        if t == 7:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Viewed '+ x+ ';\n')
        if t == 3:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Updated item in ' + x + ';\n')
        if t == 4:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Deleted item in ' + x + ';\n')
        if t == 5:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : [Program exit];\n')
        if t == 6:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Access attempt in ' + x + 'failed;\n')
        if t == 8:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Searched for ' + x + 'in database;\n')


def DisplayData(master):
    """Fetches data from database and inserts it onto tree widget"""
    cursor, conn = Database()
    cursor.execute("SELECT  Product_ID, Product,Quantity,Selling_Price FROM `stock " + year + "`")
    fetch = cursor.fetchall()
    for data in fetch:
        master.insert('', 'end', values=(data))
    cursor.close()
    conn.close()


def toggle_button():
    status = open_file("status.txt")
    print(status)
    if status == '0':
        with open("status.txt","w") as file:
            file.write('1')
        frame_color_1.grid_forget()
    elif status == '1':
        with open("status.txt","w") as file:
            file.write('0')
        frame_color_1.grid(row = 1, column = 0, sticky = tk.W+tk.N+tk.S,padx = 5, pady = 5)
        

def window_frame(master, row = 0, column = 0, rowspan = 1, width = 0, height = 0, sticky = tk.N+tk.S+tk.E+tk.W, pady =0, padx = 0, scroll = 0, anchor = tk.W, bg = None):
    """Accepts **options and returns an optionally srollable canvas window"""
    
    def on_configure(event):
        canvas.configure(scrollregion = canvas.bbox('all'))
    
    canvas = tk.Canvas(master,width = width,height =height, bg = bg)
    canvas.grid(row = row, column = column,sticky = sticky,pady = pady, rowspan = rowspan)
    
    if scroll == 1:
        scrollbar = ttk.Scrollbar(master,command = canvas.yview)
        scrollbar.grid(row = row, column = column + 1,sticky = tk.N+tk.S,pady = pady, rowspan = rowspan)
        canvas.configure(yscrollcommand = scrollbar.set)
    
    canvas.bind('<Configure>',on_configure)
    canvas_frame = tk.Frame(canvas, bg = bg)
    canvas.create_window((0,0),window = canvas_frame,anchor = anchor)
    
    return canvas_frame


def main_menu(master):
    search_var = tk.StringVar()
        
    ttk.Button(master,text = 'Search...',command = None).grid(row = 0,column = 1,sticky = tk.N+tk.E,pady = 20,padx = 2)
    ttk.Separator(master).grid(row = 1,column = 0,columnspan = 2,ipadx = 100,sticky = tk.E+tk.W,pady = 10)
    ttk.Entry(master,textvariable = search_var).grid(row = 0,column = 0,sticky = tk.N+tk.W+tk.E,pady = 20,padx = 2)
    
    def create_image_var(filename):
       """Accepts path to filename, creates and returns an image object"""
       img = Image.open(''+filename)
       image_var = ImageTk.PhotoImage(img)
       return image_var
    
    images = ["premium.png","share.png","signup.png","settings.png","question.png","privacy_policy.png"]
    images = [create_image_var(filename) for filename in images]
    
    global img1, img2, img3, img4, img5, img6
    img1, img2, img3, img4, img5, img6 = images[0], images[1], images[2], images[3], images[4], images[5]
    
    button_data = [("Get Premium",images[0],lambda : GetPremium(master),1),("Share",images[1],None,2),("Sign Up",images[2],SignUp,3),("Settings",images[3],Settings,4), ("Help",images[4],Help,5)]
    
    #create sidebar buttons
    [tk.Button(master, text = index[0], relief = tk.FLAT,image = index[1], compound = 'left',bg = 'white',height =90,font = ('Arial',10), command = index[2], anchor = tk.W).grid(row = index[3], column = 0, sticky = tk.N+tk.W+tk.S+tk.E,columnspan = 2) for index in button_data]

    
def frames(): 
    global frame_color_1, last_label, tool_bar_label
    
    frame0 = tk.Frame(root, relief = tk.FLAT)
    frame_color_1 = tk.Frame(root, relief = tk.FLAT, bg = open_file("color.txt"))
    frame2 = tk.Frame(root, relief = tk.FLAT)
    frame3 = tk.Frame(root, relief = tk.FLAT)
    
    frame0.grid(row = 0, column = 0, sticky = tk.N+tk.W+tk.E, padx = 5, pady = 5, columnspan = 2)
    frame_color_1.grid(row = 1, column = 0, sticky = tk.W+tk.N+tk.S,padx = 5, pady = 5)
    frame2.grid(row = 1, column = 1, sticky = tk.N+tk.S+tk.E+tk.W, pady = 5, padx = 5)
    frame3.grid(row = 2, column = 0, sticky = tk.N+tk.W+tk.S+tk.E,padx = 5, pady = 5,columnspan = 2)
    
    frame0.grid_columnconfigure(1,weight = 1)
    frame3.grid_columnconfigure(0,weight = 1) 
       
    #Toolbar Buttons and Label
    icons = {'exit.png' : exit_button,'question.png' : Help,'signin.png' : SignIn}
    ToolbarIcon(frame0,filename1= 'main_menu.png',button_command = toggle_button, icon_side = tk.LEFT)
    [ToolbarIcon(frame0,filename1= ''+ key,button_command = icons[key],icon_side = tk.RIGHT) for key in icons]
  
    tool_bar_label = tk.Label(frame0,text = name + ' SMS' + strftime(open_file("date.txt")), relief = tk.FLAT,font = ('Arial',12), anchor = tk.W)
    tool_bar_label.pack(side = tk.LEFT,fill = tk.X,padx = 30)
    
    #Sidebar Frame
    tk.Label(frame_color_1,text = 'Main Menu',font = ('Arial',10,'bold'),relief = tk.FLAT).grid(row = 0, column = 0, sticky = tk.N+tk.S+tk.E+tk.W, padx = 10, pady = 10)
    
    canvas_frame = window_frame(frame_color_1, row = 1, column = 0, width = 380, height = 450, scroll = 1, anchor = tk.E)
    
    main_menu(canvas_frame)
    toggle_button()
    
    last_label = tk.Label(frame_color_1,text = 'ourdigitaltimes@gmail.com',font = ('Arial',9,'bold'),bg = open_file("color.txt"),relief = tk.FLAT)
    last_label.grid(row = 2, column = 0, sticky = tk.N+tk.S+tk.E+tk.W,pady = 20)
    
    #mainframe tabs
    tabcontrol = ttk.Notebook(frame2)
    tab1 = ttk.Frame(tabcontrol)
    tab2 = ttk.Frame(tabcontrol)
    tab3 = ttk.Frame(tabcontrol)
    tab4 = ttk.Frame(tabcontrol)
    tab5 = ttk.Frame(tabcontrol)
    tab6 = ttk.Frame(tabcontrol)
    tab7 = ttk.Frame(tabcontrol)
    
    tabs = {tab1 : 'Tab 1', tab2 : ' Tab 2', tab3 : 'Tab 3', tab4 : 'Tab 4', tab5 : 'Tab 5', tab6 : 'Tab 6', tab7 : 'Tab 7'}
    [tabcontrol.add(key,text = tabs[key]) for key in tabs]
    tabcontrol.pack(expand = 1, fill = tk.BOTH)
    
    addStock(tab1)
    update(tab1)
    SalesExcel(tab2)
    showSales(tab3)
    read_txt(tab4, text = "Petty Cash & Expenses Records", filename = "Petty Cash "+year + ".txt", names = ['Date','Time','Description','Amount'])
    read_txt(tab5, text = "Depleted Stock Records", filename = month + ".txt", names = ['Date', 'Product','Cost Price'])    
    summary(tab6)
    plots(tab7)
        

    def time():
        """Configures a label to show current time at run-time"""
        string = strftime('%H:%M:%S %p') 
        lbl.config(text = 'Time: '+string) 
        lbl.after(1000, time)
    
    lbl = tk.Label(frame3)
    lbl.grid(row = 1,column = 2,sticky = tk.E)
    
    time()
    
    
def main():
    global root, style
    
    root = tk.Tk()
    root.title(name)
    root.columnconfigure(0,weight = 0)
    root.columnconfigure(1,weight = 1)
    root.attributes("-fullscreen",True)
    
    #configuring theme settings
    style = ttk.Style()
    theme = open_file("theme.txt")  
    style.theme_use(theme)
    
    #invoking frame logic and construction
    frames()
    
    root.mainloop()


def open_file(filename):
    """Accepts path of file as parameter, opens the file in read mode and returns data in the file"""
    with open(filename,"r") as file:
        data = file.read()    
    return data

def dependencies():
    dependency_filenames = {"organisation_name" : "Our Digital Times Sales Management System", "color" : "grey", "graph_style" : "ggplot", "date" : " ~ %A, %B %d, %G", "theme" : "droid", "Petty Cash " + year : "", "status" : '0'}
    for key in dependency_filenames:
        if os.path.isfile(key + ".txt"):
            pass
        else:
            with open(key + ".txt", "w") as file:
                file.write(dependency_filenames[key])
    
    
if __name__ == '__main__':
    #global variables
    global name, month, year, date
    month = strftime('%B %G')
    date = strftime('%x')
    year = strftime('%G')
    dependencies()
    name = open_file("organisation_name.txt")    
    logfile(1)
    main()